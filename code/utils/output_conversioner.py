import openpyxl


# 格式转换
# env_vertices：环境节点信息
# out_path：输出表格地址
# data_path：供应链数据文件地址
# fixed_scheme_path：不可更改方案文件地址
# model_id：模型id
# 用例：
# env = OilControlEnv(conf, sys_conf, is_simulate=True)
# observation = env.reset()
# 先reset再创建类
# conversioner = output_conversioner()
# for i in range(step):
#     state = env.obs2vec(observation)
#     all_action = agent.choose_action(state, args.train)
#     all_action = env.vec2action(all_action)
#     next_observation, reward, done, info = env.step(all_action)
#     conversioner.act2excel(env.vertices, env.signal_list, all_action)

class output_conversioner():
    def __init__(self, out_path='仿真器返回数据.xlsx', data_path='env/real_data_db',
                 fixed_scheme_path='fixed_scheme.xlsx', model_id=1):
        self.out_path = out_path
        self.output_wb = openpyxl.Workbook()
        self.ws1 = self.output_wb.active
        self.ws1.title = '运输'
        self.ws2 = self.output_wb.create_sheet('加工')
        self.ws3 = self.output_wb.create_sheet('预警')

        c = self.ws1.cell(row=1, column=1)
        c.value = 'material_code'
        c = self.ws1.cell(row=1, column=2)
        c.value = 'material_name'
        c = self.ws1.cell(row=1, column=3)
        c.value = 'from_code'
        c = self.ws1.cell(row=1, column=4)
        c.value = 'from_name'
        c = self.ws1.cell(row=1, column=5)
        c.value = 'to_code'
        c = self.ws1.cell(row=1, column=6)
        c.value = 'to_name'
        c = self.ws1.cell(row=1, column=7)
        c.value = 'mode_code'
        c = self.ws1.cell(row=1, column=8)
        c.value = 'mode_name'
        c = self.ws1.cell(row=1, column=9)
        c.value = 'period'
        c = self.ws1.cell(row=1, column=10)
        c.value = 'quantity'

        c = self.ws2.cell(row=1, column=1)
        c.value = 'node_code'
        c = self.ws2.cell(row=1, column=2)
        c.value = 'node_name'
        c = self.ws2.cell(row=1, column=3)
        c.value = 'process'
        c = self.ws2.cell(row=1, column=4)
        c.value = 'period'

        c = self.ws3.cell(row=1, column=1)
        c.value = 'node_type'
        c = self.ws3.cell(row=1, column=2)
        c.value = 'node_code'
        c = self.ws3.cell(row=1, column=3)
        c.value = 'node_name'
        c = self.ws3.cell(row=1, column=4)
        c.value = 'material_code'
        c = self.ws3.cell(row=1, column=5)
        c.value = 'signal'
        c = self.ws3.cell(row=1, column=6)
        c.value = 'storage'
        c = self.ws3.cell(row=1, column=7)
        c.value = 'upper'
        c = self.ws3.cell(row=1, column=8)
        c.value = 'lower'
        c = self.ws3.cell(row=1, column=9)
        c.value = 'period'

        self.output_wb.save(self.out_path)

        road_wb = openpyxl.load_workbook(data_path + '/s_run_transport.xlsx')
        ws = road_wb['Sheet1']
        self.road_list = []
        for i, v in enumerate(ws['A']):
            if v.value == model_id:
                road_dict = {
                    'material_code': ws['C'][i].value,
                    'material_name': ws['D'][i].value,
                    'from_code': ws['E'][i].value,
                    'from_name': ws['F'][i].value,
                    'to_code': ws['G'][i].value,
                    'to_name': ws['H'][i].value,
                    'mode_code': ws['I'][i].value,
                    'mode_name': ws['J'][i].value
                }
                self.road_list.append(road_dict)

        node_wb = openpyxl.load_workbook(data_path + '/s_info_nodes.xlsx')
        ws = node_wb['Sheet1']
        self.node_list = []
        for i, v in enumerate(ws['A']):
            if v.value == model_id:
                node_dict = {'node_code': ws['C'][i].value, 'node_name': ws['D'][i].value}
                self.node_list.append(node_dict)

        # wb = openpyxl.load_workbook(fixed_scheme_path)
        # ws = wb['运输']
        # self.fixed_scheme_list = []
        # n = 2
        # while ws['A' + str(n)].value is not None:
        #     fixed_scheme_dict = {
        #         'material_code': ws['A' + str(n)].value,
        #         'material_name': ws['B' + str(n)].value,
        #         'from_code': ws['C' + str(n)].value,
        #         'from_name': ws['D' + str(n)].value,
        #         'to_code': ws['E' + str(n)].value,
        #         'to_name': ws['F' + str(n)].value,
        #         'mode_code': ws['G' + str(n)].value,
        #         'mode_name': ws['H' + str(n)].value,
        #         'quantity': ws['J' + str(n)].value
        #     }
        #     self.fixed_scheme_list.append(fixed_scheme_dict)
        #     n += 1

        self.row1 = 2
        self.row2 = 2
        self.row3 = 2
        self.step = 0

    def reset(self):
        self.row1 = 2
        self.row2 = 2
        self.row3 = 2
        self.step = 0

    def act2excel(self, env_vertices, env_signal, in_actions):
        """
        for i in range(len(self.fixed_scheme_list)):
            c = self.ws1.cell(row=self.row1, column=1)
            c.value = self.fixed_scheme_list[i]['material_code']
            c = self.ws1.cell(row=self.row1, column=2)
            c.value = self.fixed_scheme_list[i]['material_name']
            c = self.ws1.cell(row=self.row1, column=3)
            c.value = self.fixed_scheme_list[i]['from_code']
            c = self.ws1.cell(row=self.row1, column=4)
            c.value = self.fixed_scheme_list[i]['from_name']
            c = self.ws1.cell(row=self.row1, column=5)
            c.value = self.fixed_scheme_list[i]['to_code']
            c = self.ws1.cell(row=self.row1, column=6)
            c.value = self.fixed_scheme_list[i]['to_name']
            c = self.ws1.cell(row=self.row1, column=7)
            c.value = self.fixed_scheme_list[i]['mode_code']
            c = self.ws1.cell(row=self.row1, column=8)
            c.value = self.fixed_scheme_list[i]['mode_name']
            c = self.ws1.cell(row=self.row1, column=9)
            c.value = self.step
            c = self.ws1.cell(row=self.row1, column=10)
            c.value = self.fixed_scheme_list[i]['quantity']
            self.row1 += 1
        """

        vertices = env_vertices['transfer']
        all_actions = in_actions['transfer']
        # 归一化
        for actions in all_actions.values():
            if len(actions) <= 0:
                continue
            idx_ratio = sorted(list(enumerate(actions)), key=lambda x: x[1])
            for i in range(1, len(idx_ratio)):
                actions[idx_ratio[i][0]] = idx_ratio[i][1] - idx_ratio[i - 1][1]

        for vertice in vertices.values():
            for idx, road in enumerate(vertice.nbr_road):
                material = road.material
                if material in vertice.storage.keys():  # 有的油没有对应储量，铁岭原油库
                    storage = vertice.storage[material]
                else:
                    storage = 0
                road_quantities = storage * all_actions[vertice.key][idx]

                for road_dict in self.road_list:
                    if road_dict['material_code'] == road.material and road_dict['mode_code'] == road.mode and \
                            road_dict['from_code'] == road.start and road_dict['to_code'] == road.end:
                        c = self.ws1.cell(row=self.row1, column=1)
                        c.value = road_dict['material_code']
                        c = self.ws1.cell(row=self.row1, column=2)
                        c.value = road_dict['material_name']
                        c = self.ws1.cell(row=self.row1, column=3)
                        c.value = road_dict['from_code']
                        c = self.ws1.cell(row=self.row1, column=4)
                        c.value = road_dict['from_name']
                        c = self.ws1.cell(row=self.row1, column=5)
                        c.value = road_dict['to_code']
                        c = self.ws1.cell(row=self.row1, column=6)
                        c.value = road_dict['to_name']
                        c = self.ws1.cell(row=self.row1, column=7)
                        c.value = road_dict['mode_code']
                        c = self.ws1.cell(row=self.row1, column=8)
                        c.value = road_dict['mode_name']
                        c = self.ws1.cell(row=self.row1, column=9)
                        c.value = self.step
                        c = self.ws1.cell(row=self.row1, column=10)
                        c.value = road_quantities
                        self.row1 += 1

        refinerys = env_vertices['refinery']
        for refinery in refinerys.values():
            refine_pct = in_actions['refinery'][refinery.key][0]
            process = refine_pct * (refinery.JG_upperbound[0] - refinery.JG_lowerbound[0]) + refinery.JG_lowerbound[0]
            for node_dict in self.node_list:
                if node_dict['node_code'] == refinery.key:
                    c = self.ws2.cell(row=self.row2, column=1)
                    c.value = node_dict['node_code']
                    c = self.ws2.cell(row=self.row2, column=2)
                    c.value = node_dict['node_name']
                    c = self.ws2.cell(row=self.row2, column=3)
                    c.value = process
                    c = self.ws2.cell(row=self.row2, column=4)
                    c.value = self.step
                    self.row2 += 1

        for signal_dict in env_signal:
            for node_dict in self.node_list:
                if node_dict['node_code'] == signal_dict['node_code']:
                    c = self.ws3.cell(row=self.row3, column=1)
                    c.value = signal_dict['node_type']
                    c = self.ws3.cell(row=self.row3, column=2)
                    c.value = signal_dict['node_code']
                    c = self.ws3.cell(row=self.row3, column=3)
                    c.value = node_dict['node_name']
                    c = self.ws3.cell(row=self.row3, column=4)
                    c.value = signal_dict['material_code']
                    c = self.ws3.cell(row=self.row3, column=5)
                    c.value = signal_dict['signal']
                    c = self.ws3.cell(row=self.row3, column=6)
                    c.value = signal_dict['storage']
                    c = self.ws3.cell(row=self.row3, column=7)
                    c.value = signal_dict['upper']
                    c = self.ws3.cell(row=self.row3, column=8)
                    c.value = signal_dict['lower']
                    c = self.ws3.cell(row=self.row3, column=9)
                    c.value = self.step
                    self.row3 += 1

        self.step += 1
        self.output_wb.save(self.out_path)
